from openpyxl import load_workbook
def get_data(sheet_name):
    wb = load_workbook("resource/testdata.xlsx")
    sheet = wb["Sheet1"]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data